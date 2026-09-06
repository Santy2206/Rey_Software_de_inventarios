"""
Servicio de Productos — versión BASE DE DATOS LOCAL.

Ajustado a las columnas REALES de la tabla 'productos' tal como
quedó creada en pgAdmin (distintas al diseño original en Supabase):

    productos (id, bodega_id, nombre, descripcion, sku, precio,
               stock_actual, creado_en)

Mismo contrato de siempre:
  - Cada función usa try/except
  - Siempre retorna un diccionario con 'success' y 'message'
  - Si hay datos, los incluye bajo la llave 'data'
  - SIN importaciones de Flet — solo lógica pura

productos_view.py sigue siendo un stub — cuando la construyan,
   usen estos nombres de parámetro/columna (no 'tipo', 'funcion',
   'unidad_medida' ni 'stock', que no existen en esta tabla).
"""

import os
from pathlib import Path

import openpyxl
import pandas as pd

from src.core.local_db import run_query
from src.services.auth_service import AuthService
from src.services.bitacora_service import BitacoraService


def _registrar_bitacora_productos(
    accion: str, entidad_id: str, detalle: str
):
    """Registra en bitácora sin afectar la operación principal."""
    try:
        usuario_id = AuthService.get_usuario_id()
        if usuario_id:
            BitacoraService.registrar(
                usuario_id,
                accion,
                entidad="producto",
                entidad_id=entidad_id,
                detalle=detalle,
            )
    except Exception as e:
        print(f" Error al registrar bitácora de producto: {e}")


def _aplicar_estilos_excel(ruta: str):
    """Aplica formato básico (negrita en encabezados y autoajuste)."""
    try:
        wb = openpyxl.load_workbook(ruta)
        for ws in wb.worksheets:
            header_font = openpyxl.styles.Font(bold=True)
            fill = openpyxl.styles.PatternFill(
                start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
            )
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = fill
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        value = str(cell.value) if cell.value is not None else ""
                        if len(value) > max_length:
                            max_length = len(value)
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width
        wb.save(ruta)
    except Exception as e:
        print(f" No se pudieron aplicar estilos al Excel: {e}")


class ProductosService:

    @staticmethod
    def get_all():
        """
        Trae todos los productos.
        Incluye el nombre de la bodega a la que pertenece cada uno.
        """
        print("--- Trayendo todos los productos ---")
        try:
            data = run_query("""
                SELECT p.*, b.nombre AS bodega_nombre
                FROM productos p
                LEFT JOIN bodegas b ON p.bodega_id = b.id
                ORDER BY p.nombre
                """)

            if not data:
                print("No hay productos registrados aún")
                return {"success": False, "message": "No hay productos registrados"}

            print(f"Se encontraron {len(data)} producto(s)")
            return {"success": True, "message": "Productos obtenidos", "data": data}

        except Exception as e:
            error_msg = str(e)
            print(f" Error en ProductosService.get_all: {error_msg}")
            return {
                "success": False,
                "message": f"Error al obtener productos: {error_msg}",
            }

    @staticmethod
    def get_by_bodega(bodega_id: str):
        """
        Trae todos los productos que pertenecen a una bodega específica.

        Parámetros:
            bodega_id: el id de la bodega que queremos filtrar
        """
        print(f"--- Trayendo productos de bodega: {bodega_id} ---")
        try:
            data = run_query(
                "SELECT * FROM productos WHERE bodega_id = %s ORDER BY nombre",
                (bodega_id,),
            )

            if not data:
                return {"success": False, "message": "No hay productos en esta bodega"}

            print(f"Se encontraron {len(data)} producto(s) en la bodega")
            return {"success": True, "message": "Productos obtenidos", "data": data}

        except Exception as e:
            error_msg = str(e)
            print(f" Error en ProductosService.get_by_bodega: {error_msg}")
            return {
                "success": False,
                "message": f"Error al obtener productos: {error_msg}",
            }

    @staticmethod
    def get_one(producto_id: str):
        """
        Trae un solo producto por su ID.

        Parámetros:
            producto_id: el id del producto que queremos buscar
        """
        print(f"--- Buscando producto con id: {producto_id} ---")
        try:
            producto = run_query(
                """
                SELECT p.*, b.nombre AS bodega_nombre
                FROM productos p
                LEFT JOIN bodegas b ON p.bodega_id = b.id
                WHERE p.id = %s
                """,
                (producto_id,),
                fetch_one=True,
            )

            if not producto:
                return {"success": False, "message": "Producto no encontrado"}

            print(f"Producto encontrado: {producto['nombre']}")
            return {"success": True, "message": "Producto encontrado", "data": producto}

        except Exception as e:
            error_msg = str(e)
            print(f" Error en ProductosService.get_one: {error_msg}")
            return {
                "success": False,
                "message": f"Error al buscar producto: {error_msg}",
            }

    @staticmethod
    def create(
        nombre: str,
        bodega_id: str,
        descripcion: str | None = None,
        sku: str | None = None,
        codigo: str | None = None,
        precio: float = 0.0,
        stock_actual: int = 0,
    ):
        """
        Crea un nuevo producto.

        Parámetros:
            nombre:       nombre del producto, ej: "Perfume Floral 100ml"
            bodega_id:    id de la bodega donde se almacena
            descripcion:  detalle libre del producto (opcional)
            sku:          código interno/referencia del producto (opcional)
            codigo:       código de fragancia (opcional, usado en bodegas de fragancias)
            precio:       precio de venta (por defecto 0)
            stock_actual: cantidad inicial disponible (por defecto 0)
        """
        print(f"--- Creando producto: {nombre} ---")
        try:
            producto = run_query(
                """
                INSERT INTO productos
                    (nombre, bodega_id, descripcion, sku, codigo, precio, stock_actual)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                RETURNING *
                """,
                (nombre, bodega_id, descripcion, sku, codigo, precio, stock_actual),
                fetch_one=True,
            )

            if not producto:
                return {"success": False, "message": "No se pudo crear el producto"}

            print(f" Producto '{nombre}' creado con éxito")
            return {
                "success": True,
                "message": f"Producto '{nombre}' creado con éxito",
                "data": producto,
            }

        except Exception as e:
            error_msg = str(e)
            print(f" Error en ProductosService.create: {error_msg}")
            return {
                "success": False,
                "message": f"Error al crear producto: {error_msg}",
            }

    @staticmethod
    def update(
        producto_id: str,
        nombre: str,
        bodega_id: str,
        descripcion: str | None = None,
        sku: str | None = None,
        codigo: str | None = None,
        precio: float = 0.0,
    ):
        """
        Actualiza los datos de un producto existente.
        No actualiza stock_actual aquí — el stock solo cambia a través
        de movimientos (ver movimientos_service.py).

        Parámetros:
            producto_id: el id del producto a actualizar
            nombre:      nuevo nombre
            bodega_id:   nueva bodega asignada
            descripcion: nueva descripción (opcional)
            sku:         nuevo código/referencia (opcional)
            codigo:      nuevo código de fragancia (opcional)
            precio:      nuevo precio (por defecto 0)
        """
        print(f"--- Actualizando producto id: {producto_id} ---")
        try:
            anterior = run_query(
                "SELECT precio FROM productos WHERE id = %s",
                (producto_id,),
                fetch_one=True,
            )
            precio_anterior = (
                float(anterior["precio"]) if anterior else None
            )

            producto = run_query(
                """
                UPDATE productos
                SET nombre = %s,
                    bodega_id = %s,
                    descripcion = %s,
                    sku = %s,
                    codigo = %s,
                    precio = %s
                WHERE id = %s
                RETURNING *
                """,
                (nombre, bodega_id, descripcion, sku, codigo, precio, producto_id),
                fetch_one=True,
            )

            if not producto:
                return {
                    "success": False,
                    "message": "Producto no encontrado para actualizar",
                }

            if precio_anterior is not None and float(precio or 0) != precio_anterior:
                _registrar_bitacora_productos(
                    "CAMBIO_PRECIO",
                    entidad_id=producto_id,
                    detalle=(
                        f"Precio de '{nombre}': "
                        f"${precio_anterior:,.0f} → ${float(precio or 0):,.0f}"
                    ),
                )

            print(f" Producto actualizado con éxito")
            return {
                "success": True,
                "message": "Producto actualizado con éxito",
                "data": producto,
            }

        except Exception as e:
            error_msg = str(e)
            print(f" Error en ProductosService.update: {error_msg}")
            return {
                "success": False,
                "message": f"Error al actualizar producto: {error_msg}",
            }

    @staticmethod
    def exportar_excel(bodega_id: str | None = None, ruta_salida: str | None = None):
        """
        Exporta los productos a Excel.

        Si no hay productos, genera una plantilla de importación con las
        columnas esperadas: nombre, descripcion, sku, precio, stock.

        Parámetros:
            bodega_id: filtrar por bodega; None o 'todas' trae todos.
            ruta_salida: ruta completa del archivo .xlsx. Si no se indica,
                         se guarda en assets/downloads/productos_export.xlsx.

        Retorna:
            dict: contrato estándar; incluye 'ruta', 'url' y 'es_plantilla'.
        """
        print("--- Exportando productos a Excel ---")
        try:
            project_root = Path(__file__).resolve().parents[2]
            if not ruta_salida:
                downloads_dir = project_root / "assets" / "downloads"
                downloads_dir.mkdir(parents=True, exist_ok=True)
                ruta_salida = str(downloads_dir / "productos_export.xlsx")
            else:
                ruta_salida = str(ruta_salida)

            if bodega_id and str(bodega_id).lower() not in ("", "todas"):
                res = ProductosService.get_by_bodega(bodega_id)
            else:
                res = ProductosService.get_all()

            columnas_salida = [
                "codigo",
                "nombre",
                "stock",
            ]

            if res.get("success") and res.get("data"):
                productos = res["data"]
                filas = []
                for p in productos:
                    filas.append(
                        {
                            "codigo": p.get("codigo", ""),
                            "nombre": p.get("nombre", ""),
                            "stock": int(p.get("stock_actual", 0) or 0),
                        }
                    )
                df = pd.DataFrame(filas, columns=columnas_salida)
                es_plantilla = False
            else:
                # Plantilla vacía con una fila de ejemplo aparte
                df_plantilla = pd.DataFrame(columns=columnas_salida)
                df_ejemplo = pd.DataFrame(
                    [
                        {
                            "codigo": "F-001",
                            "nombre": "Perfume Floral 100ml",
                            "stock": 15,
                        }
                    ],
                    columns=columnas_salida,
                )
                with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
                    df_plantilla.to_excel(
                        writer, sheet_name="Plantilla", index=False
                    )
                    df_ejemplo.to_excel(
                        writer, sheet_name="Ejemplo", index=False
                    )

                _aplicar_estilos_excel(ruta_salida)

                return {
                    "success": True,
                    "message": "Plantilla de productos generada",
                    "ruta": ruta_salida,
                    "url": "/downloads/productos_export.xlsx",
                    "es_plantilla": True,
                }

            df.to_excel(ruta_salida, index=False, engine="openpyxl")
            _aplicar_estilos_excel(ruta_salida)

            return {
                "success": True,
                "message": f"Exportados {len(df)} productos",
                "ruta": ruta_salida,
                "url": "/downloads/productos_export.xlsx",
                "es_plantilla": False,
            }

        except Exception as e:
            error_msg = str(e)
            print(f" Error en ProductosService.exportar_excel: {error_msg}")
            return {
                "success": False,
                "message": f"Error al exportar productos: {error_msg}",
            }

    @staticmethod
    def eliminar_todos(bodega_id: str | None = None):
        """
        Elimina todos los productos. Si se indica bodega_id, solo los de esa bodega.

        Nota: también elimina los detalles de venta y movimientos asociados
        para evitar errores de claves foráneas.
        """
        print("--- Eliminando todos los productos ---")
        try:
            from src.core.local_db import get_cursor

            with get_cursor() as cur:
                where = "WHERE bodega_id = %s" if bodega_id else ""
                params = (bodega_id,) if bodega_id else ()

                cur.execute(f"SELECT id FROM productos {where}", params)
                rows = cur.fetchall()
                ids = [r["id"] for r in rows]

                if not ids:
                    return {
                        "success": True,
                        "message": "No hay productos para eliminar",
                        "data": {"eliminados": 0},
                    }

                # Borrar dependencias que restringen la eliminación
                cur.execute(
                    "DELETE FROM venta_detalle WHERE producto_id = ANY(%s::uuid[])",
                    (ids,),
                )
                cur.execute(
                    "DELETE FROM movimientos WHERE producto_id = ANY(%s::uuid[])",
                    (ids,),
                )

                # Borrar los productos
                cur.execute(
                    "DELETE FROM productos WHERE id = ANY(%s::uuid[])",
                    (ids,),
                )
                eliminados = cur.rowcount

            _registrar_bitacora_productos(
                "ELIMINACION",
                entidad_id=None,
                detalle=f"Eliminación masiva de {eliminados} productos",
            )

            print(f" {eliminados} productos eliminados")
            return {
                "success": True,
                "message": f"Eliminados {eliminados} productos",
                "data": {"eliminados": eliminados},
            }

        except Exception as e:
            error_msg = str(e)
            print(f" Error en ProductosService.eliminar_todos: {error_msg}")
            return {
                "success": False,
                "message": f"Error al eliminar productos: {error_msg}",
            }

    @staticmethod
    def delete(producto_id: str):
        """
        Elimina un producto por su ID.

        Parámetros:
            producto_id: el id del producto a eliminar
        """
        print(f"--- Eliminando producto id: {producto_id} ---")
        try:
            eliminado = run_query(
                "DELETE FROM productos WHERE id = %s RETURNING id",
                (producto_id,),
                fetch_one=True,
            )

            if not eliminado:
                return {
                    "success": False,
                    "message": "Producto no encontrado para eliminar",
                }

            _registrar_bitacora_productos(
                "ELIMINACION",
                entidad_id=producto_id,
                detalle="Producto eliminado",
            )

            print(f" Producto eliminado con éxito")
            return {"success": True, "message": "Producto eliminado con éxito"}

        except Exception as e:
            error_msg = str(e)
            print(f" Error en ProductosService.delete: {error_msg}")
            return {
                "success": False,
                "message": f"Error al eliminar producto: {error_msg}",
            }

    @staticmethod
    def importar_excel(ruta_archivo: str, bodega_id: str):
        """
        Importa productos desde un archivo Excel (.xlsx) o CSV (.csv)
        a la bodega indicada.

        Columnas aceptadas (no importa mayúsculas/minúsculas ni tildes):
            - nombre      (o 'producto', 'name')
            - descripcion (o 'description', 'desc')
            - sku         (o 'referencia', 'ref', 'codigo_sku', 'code')
            - codigo      (o 'codigo_fragancia', 'fragancia', 'fragance_code')
            - precio      (o 'price', 'valor', 'costo')
            - stock       (o 'stock_actual', 'cantidad', 'quantity', 'qty')

        Retorna:
            dict: contrato estándar; 'data' contiene un resumen con
                  creados, errores y lista de mensajes por fila.
        """
        print(f"--- Importando productos desde: {ruta_archivo} ---")
        try:
            if not ruta_archivo or not os.path.exists(ruta_archivo):
                return {
                    "success": False,
                    "message": "No se encontró el archivo indicado",
                }

            if not bodega_id:
                return {
                    "success": False,
                    "message": "Debe seleccionar una bodega destino",
                }

            extension = os.path.splitext(ruta_archivo)[1].lower()
            if extension == ".csv":
                df = pd.read_csv(ruta_archivo, dtype=str)
            elif extension in (".xlsx", ".xls"):
                df = pd.read_excel(ruta_archivo, dtype=str)
            else:
                return {
                    "success": False,
                    "message": "Formato no soportado. Use .xlsx o .csv",
                }

            if df.empty:
                return {
                    "success": False,
                    "message": "El archivo está vacío o no tiene datos",
                }

            # Normalizar nombres de columnas
            def _normalizar(col: str) -> str:
                return (
                    col.lower()
                    .strip()
                    .replace("á", "a")
                    .replace("é", "e")
                    .replace("í", "i")
                    .replace("ó", "o")
                    .replace("ú", "u")
                    .replace("ñ", "n")
                )

            columnas = {_normalizar(c): c for c in df.columns}

            mapeos = {
                "nombre": ["nombre", "producto", "name"],
                "descripcion": ["descripcion", "description", "desc"],
                "sku": ["sku", "referencia", "ref", "codigo_sku", "code"],
                "codigo": [
                    "codigo",
                    "codigo_fragancia",
                    "fragancia",
                    "fragance_code",
                    "fragance",
                ],
                "precio": ["precio", "price", "valor", "costo"],
                "stock_actual": [
                    "stock",
                    "stock_actual",
                    "cantidad",
                    "quantity",
                    "qty",
                ],
            }

            def _buscar_columna(nombres):
                for n in nombres:
                    if n in columnas:
                        return columnas[n]
                return None

            col_nombre = _buscar_columna(mapeos["nombre"])
            col_descripcion = _buscar_columna(mapeos["descripcion"])
            col_sku = _buscar_columna(mapeos["sku"])
            col_codigo = _buscar_columna(mapeos["codigo"])
            col_precio = _buscar_columna(mapeos["precio"])
            col_stock = _buscar_columna(mapeos["stock_actual"])

            if not col_nombre:
                return {
                    "success": False,
                    "message": "No se encontró una columna de nombre en el archivo",
                }

            creados = 0
            errores = 0
            detalle = []

            for _, fila in df.iterrows():
                nombre = (fila.get(col_nombre) or "").strip()
                if not nombre:
                    errores += 1
                    detalle.append("Fila con nombre vacío omitida")
                    continue

                sku_raw = fila.get(col_sku) if col_sku else None
                sku = (str(sku_raw).strip() if pd.notna(sku_raw) else "")

                descripcion_raw = fila.get(col_descripcion) if col_descripcion else None
                descripcion = (
                    str(descripcion_raw).strip()
                    if pd.notna(descripcion_raw)
                    else None
                )

                codigo_raw = fila.get(col_codigo) if col_codigo else None
                codigo = (
                    str(codigo_raw).strip()
                    if pd.notna(codigo_raw)
                    else None
                )

                precio_raw = fila.get(col_precio) if col_precio else 0
                try:
                    precio = float(str(precio_raw).replace(",", "").strip() or 0)
                except ValueError:
                    precio = 0.0

                stock_raw = fila.get(col_stock) if col_stock else 0
                try:
                    stock = int(str(stock_raw).replace(",", "").strip() or 0)
                except ValueError:
                    stock = 0

                resultado = ProductosService.create(
                    nombre=nombre,
                    bodega_id=bodega_id,
                    descripcion=descripcion,
                    sku=sku,
                    codigo=codigo,
                    precio=precio,
                    stock_actual=stock,
                )

                if resultado.get("success"):
                    creados += 1
                    detalle.append(f"Creado: {nombre}")
                else:
                    errores += 1
                    detalle.append(f"Error en '{nombre}': {resultado.get('message')}")

            return {
                "success": True,
                "message": (
                    f"Importación finalizada: {creados} creados, "
                    f"{errores} errores"
                ),
                "data": {
                    "creados": creados,
                    "errores": errores,
                    "detalle": detalle,
                },
            }

        except Exception as e:
            error_msg = str(e)
            print(f" Error en ProductosService.importar_excel: {error_msg}")
            return {
                "success": False,
                "message": f"Error al importar productos: {error_msg}",
            }
